import os
import json
import datetime
import openpyxl

# Configuration
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Data')
TMP_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.tmp')
OUTPUT_FILE = os.path.join(TMP_DIR, 'debug_data.json')

def parse_header(ws):
    """
    Validates Row 1 and Row 2 headers.
    Returns True if valid, raises error otherwise.
    """
    # check Row 1: A1=Date, C1=Total Marks
    r1 = [c.value for c in ws[1]]
    if len(r1) < 4:
        raise ValueError("Row 1 must have at least 4 columns")
    
    if r1[0] != 'Date':
        raise ValueError(f"Cell A1 should be 'Date', found '{r1[0]}'")
    if r1[2] != 'Total Marks':
        raise ValueError(f"Cell C1 should be 'Total Marks', found '{r1[2]}'")
        
    # check Row 2: A2=Name, B2=Marks
    r2 = [c.value for c in ws[2]]
    if r2[0] != 'Name':
        raise ValueError(f"Cell A2 should be 'Name', found '{r2[0]}'")
    if r2[1] != 'Marks':
        raise ValueError(f"Cell B2 should be 'Marks', found '{r2[1]}'")
        
    return {
        'date': r1[1],
        'total_marks': r1[3]
    }

def parse_sheet(ws, sheet_name):
    try:
        meta = parse_header(ws)
    except Exception as e:
        print(f"  [WARN] Skipping sheet '{sheet_name}': {e}")
        return None

    students = []
    # Iterate from row 3
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]: # Stop at empty row
            continue
            
        name = row[0]
        raw_marks = row[1]
        comments = row[2] if len(row) > 2 else None
        
        # Handle AB/ABS
        marks = None
        if raw_marks in ['AB', 'ABS', 'ab', 'abs']:
            marks = None
        elif isinstance(raw_marks, (int, float)):
            marks = float(raw_marks)
        else:
            # Try to convert string to number, otherwise None
            try:
                marks = float(raw_marks)
            except:
                marks = None
        
        students.append({
            'name': name,
            'marks': marks,
            'comments': comments
        })
        
    return {
        'topicName': sheet_name,
        'date': meta['date'].isoformat() if isinstance(meta['date'], datetime.date) else str(meta['date']),
        'totalMarks': meta['total_marks'],
        'students': students
    }

def main():
    if not os.path.exists(TMP_DIR):
        os.makedirs(TMP_DIR)
        
    print(f"Scanning Data Directory: {DATA_DIR}")
    
    all_data = {}
    
    # Traverse Data directory
    for class_name in os.listdir(DATA_DIR):
        class_path = os.path.join(DATA_DIR, class_name)
        if not os.path.isdir(class_path) or class_name.startswith('.'):
            continue
            
        all_data[class_name] = {}
            
        for subject_name in os.listdir(class_path):
            subject_path = os.path.join(class_path, subject_name)
            if not os.path.isdir(subject_path):
                continue
                
            subject_data = {
                'subjectName': subject_name,
                'className': class_name,
                'topics': []
            }
            
            for file in os.listdir(subject_path):
                if file.endswith('.xlsx') and not file.startswith('~$'):
                    full_path = os.path.join(subject_path, file)
                    topic_name = os.path.splitext(file)[0]
                    
                    print(f"Processing: {class_name} / {subject_name} / {topic_name}")
                    
                    try:
                        wb = openpyxl.load_workbook(full_path, data_only=True)
                        if wb.sheetnames:
                            sheet_name = wb.sheetnames[0]
                            topic = parse_sheet(wb[sheet_name], sheet_name or topic_name)
                            if topic:
                                subject_data['topics'].append(topic)
                    except Exception as e:
                        print(f"  [ERROR] Failed to read {file}: {e}")
                        
            all_data[class_name][subject_name] = subject_data

    # Output results
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(all_data, f, indent=2)
        
    print(f"Done. Parsed data saved to {OUTPUT_FILE}")

if __name__ == '__main__':
    main()
